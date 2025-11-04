import streamlit as st
import pandas as pd
from pathlib import Path
import zipfile
import io
from openpyxl import load_workbook

st.set_page_config(
    page_title="Excel Sheet Splitter",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Excel Sheet Splitter")
st.markdown("Upload an Excel file with multiple sheets, and we'll split each sheet into a separate Excel file.")

# Initialize session state
if 'processed_files' not in st.session_state:
    st.session_state.processed_files = {}
if 'temp_file_data' not in st.session_state:
    st.session_state.temp_file_data = None
if 'sheet_names' not in st.session_state:
    st.session_state.sheet_names = []
if 'original_filename' not in st.session_state:
    st.session_state.original_filename = ""

# File uploader
uploaded_file = st.file_uploader(
    "Choose an Excel file",
    type=['xlsx', 'xls'],
    help="Upload an Excel file containing multiple sheets",
    key="file_uploader"
)

# If a new file is uploaded, process it
if uploaded_file is not None:
    try:
        # Check if this is a new file (different from what we have in session state)
        current_file_id = f"{uploaded_file.name}_{uploaded_file.size}"
        
        # Save uploaded file temporarily to read with openpyxl
        temp_file = uploaded_file.read()
        
        # Read all sheets from the Excel file using pandas for display
        excel_file = pd.ExcelFile(io.BytesIO(temp_file))
        sheet_names = excel_file.sheet_names
        
        # Store in session state
        st.session_state.temp_file_data = temp_file
        st.session_state.sheet_names = sheet_names
        st.session_state.original_filename = Path(uploaded_file.name).stem
        
        st.success(f"✅ File loaded successfully! Found {len(sheet_names)} sheet(s).")
        
        # Display sheet names
        st.subheader("📋 Sheets Found:")
        for idx, sheet_name in enumerate(sheet_names, 1):
            st.write(f"{idx}. {sheet_name}")
        
        # Process button
        if st.button("🔄 Split Sheets into Separate Files", type="primary"):
            # Get data from session state
            temp_file = st.session_state.temp_file_data
            sheet_names = st.session_state.sheet_names
            original_filename = st.session_state.original_filename
            
            # Progress bar
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            split_files_data = {}
            
            # Process each sheet
            for idx, sheet_name in enumerate(sheet_names):
                status_text.text(f"Processing sheet: {sheet_name} (preserving all formatting)...")
                
                # BEST METHOD: Load the full workbook and remove other sheets
                # This preserves 100% of formatting because we're using the original file structure
                wb_new = load_workbook(io.BytesIO(temp_file), keep_links=False, data_only=False)
                
                # Remove all sheets except the one we want
                # This way ALL formatting is preserved (colors, fonts, borders, images, charts, etc.)
                sheets_to_remove = [name for name in wb_new.sheetnames if name != sheet_name]
                for sheet_to_remove in sheets_to_remove:
                    wb_new.remove(wb_new[sheet_to_remove])
                
                # Ensure the remaining sheet has the correct name
                if sheet_name in wb_new.sheetnames:
                    wb_new[sheet_name].title = sheet_name
                
                # Clean sheet name for filename (remove invalid characters)
                clean_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
                clean_sheet_name = clean_sheet_name.replace(' ', '_')
                
                # Create output filename
                output_filename = f"{clean_sheet_name}.xlsx"
                
                # Save to memory buffer instead of file system
                output_buffer = io.BytesIO()
                wb_new.save(output_buffer)
                output_buffer.seek(0)
                
                # Store in session state
                split_files_data[output_filename] = output_buffer.getvalue()
                
                wb_new.close()
                
                # Update progress
                progress_bar.progress((idx + 1) / len(sheet_names))
            
            # Store processed files in session state
            st.session_state.processed_files = split_files_data
            
            status_text.text("✅ All sheets processed successfully!")
            st.rerun()  # Rerun to show download buttons
            
    except Exception as e:
        st.error(f"❌ An error occurred: {str(e)}")
        st.exception(e)

# Display download section if files are processed
if st.session_state.processed_files:
    st.success("🎉 Files created successfully! You can download them anytime below.")
    
    # Create a zip file containing all split files
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for filename, file_data in st.session_state.processed_files.items():
            zip_file.writestr(filename, file_data)
    
    zip_buffer.seek(0)
    
    # Display download button for ZIP
    st.subheader("📥 Download All Files as ZIP")
    st.download_button(
        label="📦 Download All Files as ZIP",
        data=zip_buffer.getvalue(),
        file_name=f"{st.session_state.original_filename}_split_files.zip",
        mime="application/zip",
        key="download_zip"
    )
    
    # Display individual download buttons
    st.subheader("📥 Download Individual Files:")
    
    for filename, file_data in st.session_state.processed_files.items():
        st.download_button(
            label=f"📄 Download {filename}",
            data=file_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{filename}"
        )

# Instructions
with st.expander("ℹ️ How to use"):
    st.markdown("""
    1. **Upload an Excel file** that contains multiple sheets
    2. The app will display all sheet names found in the file
    3. Click **"Split Sheets into Separate Files"** button
    4. Download the ZIP file containing all split files, or download individual files
    5. Each sheet will be saved as a separate Excel file with the sheet name as the filename
    
    **Example:**
    - Original file: `data.xlsx` with sheets: `nov`, `oc`, `p`, `c`
    - Output files: `nov.xlsx`, `oc.xlsx`, `p.xlsx`, `c.xlsx`
    """)

